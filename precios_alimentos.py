# -*- coding: utf-8 -*-
"""Precios de los alimentos y de la leche, desde una planilla que carga el tambo.

POR QUÉ EXISTE ESTE MÓDULO. Los 70 ingredientes de Haasten tienen `price: 0` —
nadie los cargó del lado del mixer— y La Serenísima publica solo datos físicos,
sin importes. O sea que los dos precios que hacen falta para llegar a la plata
no están en ningún sistema conectado. Los carga el tambo, en un Excel, y de acá
los lee la app. Es lo que desbloquea el costo de alimentación y los litros
libres; la eficiencia de conversión no los necesita porque es física.

CÓMO ESTÁ ARMADA LA PLANILLA (`Precios_alimentos.xlsx`, generada por la app):

  * Los insumos van en USD POR TONELADA de materia FRESCA, no en pesos. El
    grano se comercializa así en Argentina y es la cifra que no se vence; el
    peso se mueve con el tipo de cambio. Hay UNA celda de tipo de cambio y
    todos los $/kg salen de ahí.
  * Una columna "$/kg manual" PISA el cálculo en dólares, para el insumo que se
    compró a precio fijo en pesos.
  * El precio de la leche está en la misma planilla, en USD/litro, con su
    "$/litro manual" al lado.
  * Cada fila dice si su precio es ESTIMADO. Los que genera la app arrancan
    estimados y `resumen()` lo reporta, para que ninguna pantalla muestre un
    costo inventado como si fuera el del tambo — la misma regla que se aplica a
    los datos genéticos simulados (ver genetica.py).

SE BUSCA POR ETIQUETA, NO POR POSICIÓN. El tambo va a abrir este archivo y
agregar filas, notas, una columna. Si el lector dependiera de que el tipo de
cambio está en B4, cualquier fila insertada arriba lo rompería en silencio —o
peor, leería otro número como si fuera el tipo de cambio. Así que se busca el
encabezado de la tabla y las etiquetas de las celdas sueltas por su texto.

EL PRECIO DE LA LECHE ES EL EFECTIVO POR LITRO. En Argentina la leche se
liquida por kg de sólidos con bonificaciones, no por litro: el número que va en
la planilla es lo cobrado dividido por los litros entregados. Se aclara en el
Instructivo de la planilla, porque de otro modo alguien pone ahí un precio de
lista y todos los litros libres salen corridos.
"""
import os
import threading
import unicodedata

RUTA_DEFECTO = os.path.join(os.path.dirname(__file__), "Precios_alimentos.xlsx")

HOJA = "Precios"

# Etiquetas que se buscan para las dos celdas sueltas. Se comparan normalizadas
# y por "empieza con", así que agregarles texto no las rompe.
ETIQUETA_TC = "tipo de cambio"
ETIQUETA_LECHE = "precio de la leche"
ETIQUETA_LECHE_MANUAL = "$/litro manual"

# Encabezados de la tabla de insumos. La clave del dict es el nombre interno.
COLUMNAS = {
    "ingrediente": "ingrediente",
    "usd_tn": "usd/tonelada",
    "precio_kg_manual": "$/kg manual",
    "estimado": "estimado",
    "notas": "notas",
}

_lock = threading.Lock()
_cache = {"clave": None, "datos": None}


def _norm(texto) -> str:
    """Sin acentos, sin mayúsculas, sin espacios de más. Mismo criterio que
    `genetica._norm`: los nombres de ingrediente vienen del mixer escritos de
    cualquier manera ('EXPELLER SOJA' y 'ExpelerSoja' conviven en la misma
    cuenta), así que el cruce no puede depender de cómo se tipeó."""
    if texto is None:
        return ""
    s = unicodedata.normalize("NFKD", str(texto).strip().lower())
    return "".join(c for c in s if not unicodedata.combining(c))


def _numero(v):
    """Número o None. Nunca 0 por defecto: un cero se propaga como un precio
    real y miente (la misma razón por la que `haasten.ingredientes()` traduce
    su `price: 0` a None). El 0 explícito de AGUA y SOBRANTE sí es un precio y
    se respeta — la diferencia es que ahí lo escribió alguien."""
    if v is None or v == "":
        return None
    if isinstance(v, str):
        # Una fórmula sin recalcular llega como texto '=D8*$B$4/1000': no es un
        # número y no hay que adivinarlo.
        v = v.replace("$", "").replace(",", "").strip()
        if v.startswith("="):
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _celda_por_etiqueta(ws, etiqueta: str, max_fila: int = 30):
    """Valor de la celda que sigue a una etiqueta de texto, buscando en las
    primeras filas. Devuelve (valor, fila) o (None, None)."""
    objetivo = _norm(etiqueta)
    for fila in ws.iter_rows(min_row=1, max_row=max_fila):
        for i, c in enumerate(fila):
            if _norm(c.value).startswith(objetivo):
                # El valor está en la primera celda a la derecha que tenga algo.
                for siguiente in fila[i + 1:]:
                    n = _numero(siguiente.value)
                    if n is not None:
                        return n, c.row
                return None, c.row
    return None, None


def _leer_excel(ruta: str) -> dict:
    import openpyxl   # import perezoso: si falta, `leer()` lo reporta
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[HOJA] if HOJA in wb.sheetnames else wb[wb.sheetnames[0]]

    tc, _ = _celda_por_etiqueta(ws, ETIQUETA_TC)
    leche_usd, fila_leche = _celda_por_etiqueta(ws, ETIQUETA_LECHE)
    leche_manual, _ = _celda_por_etiqueta(ws, ETIQUETA_LECHE_MANUAL)

    # Encabezado de la tabla: la primera fila que tenga la columna del nombre.
    objetivo = _norm(COLUMNAS["ingrediente"])
    pos, fila_cab = None, None
    for fila in ws.iter_rows(min_row=1, max_row=60):
        celdas = {_norm(c.value): c.column for c in fila if c.value is not None}
        if objetivo in celdas:
            pos = {clave: celdas.get(_norm(nombre)) for clave, nombre in COLUMNAS.items()}
            fila_cab = fila[0].row
            break
    if pos is None:
        wb.close()
        return {"error": f"No se encontró la columna «{COLUMNAS['ingrediente']}» en {os.path.basename(ruta)}."}

    precios = {}
    for fila in ws.iter_rows(min_row=fila_cab + 1):
        def celda(clave):
            col = pos.get(clave)
            return fila[col - 1].value if col and col <= len(fila) else None

        nombre = celda("ingrediente")
        if not nombre or not str(nombre).strip():
            continue
        nombre = str(nombre).strip()
        # Las filas de sección ("— SE USAN HOY (18 ingredientes...) —") no son
        # ingredientes. Se saltean por la marca, no por su posición.
        if nombre.startswith("—") or nombre.startswith("--"):
            continue
        usd_tn = _numero(celda("usd_tn"))
        manual = _numero(celda("precio_kg_manual"))
        if manual is not None:
            precio_kg, via = manual, "manual"
        elif usd_tn is not None and tc:
            precio_kg, via = usd_tn * tc / 1000.0, "usd"
        else:
            # Sin precio: la fila existe pero no aporta. No entra al dict, así
            # `cobertura()` la cuenta como faltante en vez de como cero.
            continue
        precios[_norm(nombre)] = {
            "nombre": nombre,
            "precio_kg": precio_kg,
            "usd_tn": usd_tn,
            "via": via,
            "estimado": _norm(celda("estimado")) in ("si", "sí", "true", "1"),
            "notas": (str(celda("notas")).strip() or None) if celda("notas") else None,
        }
    wb.close()

    if leche_manual is not None:
        leche_kg, leche_via = leche_manual, "manual"
    elif leche_usd is not None and tc:
        leche_kg, leche_via = leche_usd * tc, "usd"
    else:
        leche_kg, leche_via = None, None

    return {
        "error": None,
        "precios": precios,
        "tipo_cambio": tc,
        "leche_usd_litro": leche_usd,
        "precio_litro": leche_kg,
        "leche_via": leche_via,
        "archivo": os.path.basename(ruta),
    }


def leer(ruta: str = None) -> dict:
    """Precios de la planilla, releídos solo si cambió el archivo.

    `ruta` None = la de por defecto en el directorio de la app. La config del
    tambo puede apuntar a otra (ver configuracion_tambo.ruta_precios).
    """
    ruta = ruta or RUTA_DEFECTO
    with _lock:
        if not os.path.exists(ruta):
            return {"error": f"No se encontró la planilla de precios ({ruta}).",
                    "precios": {}, "precio_litro": None, "ruta": ruta}
        clave = (ruta, os.path.getmtime(ruta))
        if _cache["clave"] == clave and _cache["datos"] is not None:
            return _cache["datos"]
        try:
            datos = _leer_excel(ruta)
        except Exception as exc:  # noqa: BLE001
            datos = {"error": f"No se pudo leer {os.path.basename(ruta)}: {exc}",
                     "precios": {}, "precio_litro": None}
        datos["ruta"] = ruta
        datos.setdefault("precios", {})
        datos.setdefault("precio_litro", None)
        _cache["clave"] = clave
        _cache["datos"] = datos
        return datos


def de(nombre, datos: dict = None) -> dict | None:
    """Precio de un ingrediente por su nombre, como lo escribe el mixer."""
    d = datos if datos is not None else leer()
    return (d.get("precios") or {}).get(_norm(nombre))


def resumen(ruta: str = None) -> dict:
    """Estado de la planilla para la pantalla: cuántos precios hay, cuántos son
    ESTIMADOS y si falta el precio de la leche.

    El aviso de estimados no es cosmético: sin él una pantalla de costos muestra
    plata inventada con la misma cara que la real, y alguien decide con eso.
    """
    d = leer(ruta)
    precios = d.get("precios") or {}
    estimados = [p for p in precios.values() if p.get("estimado")]
    faltan = []
    if not d.get("precio_litro"):
        faltan.append("el precio de la leche")
    if not precios:
        faltan.append("los precios de los insumos")
    aviso = None
    if estimados:
        aviso = (f"Hay precios ESTIMADOS en uso ({len(estimados)} de {len(precios)} insumos): "
                 f"sirven para ver el cálculo, no para decidir. Cargá los precios del tambo "
                 f"en la planilla y borrá el «SI» de la columna Estimado.")
    return {
        "error": d.get("error"),
        "ruta": d.get("ruta"),
        "archivo": d.get("archivo"),
        "insumos": len(precios),
        "estimados": len(estimados),
        "hay_estimados": bool(estimados),
        "aviso_estimado": aviso,
        "tipo_cambio": d.get("tipo_cambio"),
        "precio_litro": d.get("precio_litro"),
        "leche_usd_litro": d.get("leche_usd_litro"),
        "falta": faltan or None,
    }

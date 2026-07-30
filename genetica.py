# -*- coding: utf-8 -*-
"""Riesgo genético de enfermedad por toro, leído del catálogo `Toros.xlsx`.

QUÉ ES ESTE NÚMERO, Y QUÉ NO ES
===============================
Devuelve un `riesgo` de 0 a 100 por toro. Es un RANKING RELATIVO entre los
toros del propio catálogo: 100 es el peor de la lista para rasgos de salud, 0
el mejor. NO es la probabilidad de que la vaca se enferme, y la pantalla no
debe presentarlo así. Tres razones concretas:

  * Las columnas del catálogo son PTAs: valores relativos a la población de
    referencia de la evaluación genética, no riesgos absolutos.
  * La heredabilidad de estos rasgos es BAJA (mastitis ~0,03-0,12; células
    somáticas ~0,10-0,15 según la bibliografía). La genética explica una
    fracción chica de quién se enferma; el resto es ambiente y manejo.
  * Sólo se conoce el PADRE. La vaca hereda mitad del padre y mitad de la
    madre, así que esto cubre a lo sumo la mitad de su genética.

Por eso en el índice de salud entra con PESO BAJO y como CONTEXTO/desempate,
nunca como evidencia de que la vaca esté enferma HOY: la genética no cambia
día a día, y si pesara fuerte las mismas vacas encabezarían la lista para
siempre — que es justo lo contrario de para qué sirve "Atención vacas".

EL RANKING SE REESCALA AL AGREGAR TOROS. Al ser percentil dentro del catálogo,
sumar toros mueve los valores de los que ya estaban. Es el precio de no
inventar rangos de referencia absolutos (serían específicos de raza y
evaluación, y no los tenemos). Se avisa en `escala` para que la pantalla lo
pueda decir.

RASGOS QUE USA, con su dirección (verificada contra los rangos del archivo):

    Cél. Somáticas   SCS   ~2,5-3,2   MÁS BAJO ES MEJOR  -> se invierte
    Mastitis         PTA de resistencia a mastitis        MÁS ALTO ES MEJOR
    Health Index     índice compuesto de salud            MÁS ALTO ES MEJOR
    Vida Productiva  PL, en meses                         MÁS ALTO ES MEJOR

`SCE`/`SSB`/`DSB` (facilidad de parto y mortinatos) quedan AFUERA a propósito:
son otro dominio (parto, no enfermedad de la vaca en ordeñe) y no está
confirmado en qué escala vienen en este archivo — meterlos sin saber el signo
daría un índice invertido, que es peor que no tenerlos.

Un toro que no está en el catálogo, o al que le faltan todos los rasgos, sale
con `riesgo = None` y `motivo`. NUNCA sale con riesgo 0: "no sé" y "el mejor
de la lista" no son lo mismo.
"""
import os
import threading
import unicodedata

_DIR = os.path.dirname(__file__)

# Catálogos que se leen y se combinan, en orden. Todos opcionales: falta uno y
# se sigue con los demás.
#   Toros.xlsx                     catálogo del proveedor de genética (toros
#                                  para USAR a futuro; hoy no incluye a los
#                                  padres del rodeo actual)
#   Padres_del_rodeo.xlsx          los padres REALES del rodeo (`FatherId` de
#                                  PedigreeInfo), para completar a mano — es el
#                                  que hace falta para puntuar las vacas de hoy
#   Padres_del_rodeo_SIMULADO.xlsx datos FICTICIOS para probar el pipeline
#
# El tercero lleva una columna "Simulado" en SI: los toros que salgan de ahí
# quedan marcados `simulado: True` y `resumen()` lo reporta, para que ninguna
# pantalla muestre un riesgo inventado sin decirlo. Si un toro aparece en más
# de un archivo gana el PRIMERO de la lista (el dato real le gana al simulado).
RUTAS = [os.path.join(_DIR, n) for n in
         ("Toros.xlsx", "Padres_del_rodeo.xlsx", "Padres_del_rodeo_SIMULADO.xlsx")]

COL_SIMULADO = "simulado"

# Cómo se explica cada rasgo en pantalla: nombre en criollo y qué significa el
# valor. La dirección (si más alto es mejor) sale de RASGOS, más abajo — acá
# solo el texto, para que la tarjeta pueda decir POR QUÉ un toro puntúa mal en
# vez de mostrar un percentil suelto.
RASGOS_LABEL = {
    "scs": ("Células somáticas (SCS)", "más bajo es mejor"),
    "mastitis": ("Resistencia a mastitis", "más alto es mejor"),
    "health_index": ("Índice de salud", "más alto es mejor"),
    "vida_productiva": ("Vida productiva (meses)", "más alto es mejor"),
}


def explicar(toro: dict | None) -> list:
    """[{clave, label, valor, direccion, percentil, texto}] por rasgo del toro.

    `percentil` es el aporte AL RIESGO (0 = el mejor del catálogo, 100 = el
    peor), ya con la dirección aplicada — así todas las barras se leen igual:
    más largo es peor. El `valor` es el número crudo del catálogo, que es el
    que el asesor genético reconoce."""
    if not toro:
        return []
    out = []
    for clave, (label, direccion) in RASGOS_LABEL.items():
        if clave not in toro.get("percentiles", {}):
            continue
        pct = toro["percentiles"][clave]
        crudo = toro["rasgos"].get(clave)
        if pct >= 75:
            juicio = f"entre los peores del catálogo ({round(pct)} de 100)"
        elif pct >= 50:
            juicio = f"algo por debajo del promedio ({round(pct)} de 100)"
        elif pct >= 25:
            juicio = f"algo mejor que el promedio ({round(pct)} de 100)"
        else:
            juicio = f"entre los mejores del catálogo ({round(pct)} de 100)"
        out.append({
            "clave": clave, "label": label, "valor": crudo,
            "direccion": direccion, "percentil": round(pct, 1),
            # `texto` se lee solo (lleva el label adentro); `detalle` es el mismo
            # sin el label, para quien ya lo muestra aparte — sin esto la ficha
            # imprimía "Mastitis / Mastitis: 3.3 ...". Mismo par de campos que
            # devuelve `herencia.explicar_madre`, así el frontend no ramifica.
            "detalle": f"{crudo} ({direccion}) — {juicio}",
            "texto": f"{label}: {crudo} ({direccion}) — {juicio}",
        })
    return out

# Nombre de la columna en el Excel -> (clave interna, más_alto_es_mejor)
RASGOS = {
    "cel. somaticas": ("scs", False),
    "mastitis": ("mastitis", True),
    "health index": ("health_index", True),
    "vida productiva": ("vida_productiva", True),
}

# Columnas de identificación del toro (para poder cruzarlo con la base).
COL_NAAB = "naab code"
COL_NOMBRE = "nombre"
COL_REG = "reg name"

# Pedigrí del propio toro, para poder mostrar el árbol en la ficha del animal:
# su padre (abuelo paterno de la vaca) y su abuelo materno.
COL_SIRE = "sire"
COL_MGS = "mgs"

# Rasgos de PRODUCCIÓN. Van aparte de los de salud (RASGOS) porque responden
# otra pregunta —"¿va a producir?" contra "¿se va a enfermar?"— y no entran a
# ningún índice de riesgo: se muestran como contexto en la ficha del animal.
# En todos, más alto es mejor.
RASGOS_PRODUCCION = {
    "pta milk": ("pta_leche", "Leche (PTA, lb)"),
    "pta fat": ("pta_grasa", "Grasa (PTA, lb)"),
    "pta pro": ("pta_proteina", "Proteína (PTA, lb)"),
    "tpi": ("tpi", "TPI (índice general)"),
    "merito neto": ("merito_neto", "Mérito Neto (US$)"),
}

# {(archivo, mtime)...: catálogo}. Indexado por el juego de archivos y no por un
# slot único, porque dos tambos pueden apuntar a catálogos distintos.
_cache = {}
_lock = threading.Lock()


def _norm(texto) -> str:
    """Encabezado normalizado: sin acentos, minúsculas, espacios colapsados.
    El archivo viene del proveedor de genética con acentos y dobles espacios
    ('Cél. Somáticas', hoja 'Spanish - Holandés  List'), así que comparar el
    texto crudo es frágil."""
    if texto is None:
        return ""
    s = unicodedata.normalize("NFKD", str(texto))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def _percentil(valor, ordenados: list) -> float:
    """Posición de `valor` dentro de `ordenados` (ascendente), 0-100.
    Con un solo toro no hay ranking posible y devuelve 50 (el medio): decir
    "es el peor" o "el mejor" con N=1 sería inventar."""
    n = len(ordenados)
    if n <= 1:
        return 50.0
    menores = sum(1 for v in ordenados if v < valor)
    iguales = sum(1 for v in ordenados if v == valor)
    # Percentil medio del empate, para que dos toros idénticos den lo mismo.
    return 100.0 * (menores + (iguales - 1) / 2.0) / (n - 1)


def _leer_excel(ruta: str) -> dict:
    """{clave_toro: {naab, nombre, reg, rasgos:{...}}} — clave = NAAB en
    mayúsculas, y además se indexa por nombre para poder cruzar cuando la base
    guarda el nombre del padre y no el código.

    El encabezado no siempre está en la primera fila (el archivo simulado abre
    con una fila de aviso), así que se busca la primera fila que contenga la
    columna de nombre del toro."""
    import openpyxl   # import perezoso: si falta, `catalogo()` lo reporta
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = ws.iter_rows(values_only=True)
    pos = None
    for cruda in filas:
        encabezado = [_norm(c) for c in cruda]
        if COL_NOMBRE in encabezado:
            pos = {nombre: i for i, nombre in enumerate(encabezado)}
            break
    if pos is None:
        return {}
    toros = {}
    for fila in filas:
        def celda(nombre):
            i = pos.get(nombre)
            return fila[i] if i is not None and i < len(fila) else None

        naab = celda(COL_NAAB)
        nombre = celda(COL_NOMBRE)
        if not naab and not nombre:
            continue          # fila vacía o de relleno
        rasgos = {}
        for col, (clave, _) in RASGOS.items():
            v = celda(col)
            if isinstance(v, (int, float)):
                rasgos[clave] = float(v)
        produccion = {}
        for col, (clave, _) in RASGOS_PRODUCCION.items():
            v = celda(col)
            if isinstance(v, (int, float)):
                produccion[clave] = float(v)
        toros[str(naab or nombre).strip().upper()] = {
            "produccion": produccion,
            "sire": (str(celda(COL_SIRE)).strip() or None) if celda(COL_SIRE) else None,
            "mgs": (str(celda(COL_MGS)).strip() or None) if celda(COL_MGS) else None,
            "naab": str(naab).strip() if naab else None,
            "nombre": str(nombre).strip() if nombre else None,
            "reg": str(celda(COL_REG)).strip() if celda(COL_REG) else None,
            "rasgos": rasgos,
            "simulado": _norm(celda(COL_SIMULADO)) in ("si", "sí", "true", "1"),
            "archivo": os.path.basename(ruta),
        }
    wb.close()
    return toros


def catalogo(rutas=None) -> dict:
    """{toros: {clave: {...riesgo, percentiles, rasgos...}}, por_nombre: {},
    escala: str, error: str|None}. Se relee solo si cambió el archivo.

    `rutas` None = los archivos por defecto del directorio de la app (`RUTAS`).
    La config del tambo puede apuntar a otra carpeta — ver
    `configuracion_tambo.rutas_toros`, que es quien resuelve archivo/carpeta.
    """
    with _lock:
        candidatas = list(rutas) if rutas else RUTAS
        presentes = [r for r in candidatas if os.path.exists(r)]
        if not presentes:
            return {"toros": {}, "por_nombre": {}, "error":
                    "No se encontró ningún catálogo de toros ("
                    + ", ".join(os.path.basename(r) for r in candidatas) + ")."}
        # La clave de caché es la lista de (archivo, mtime): cambiar cualquiera
        # de los archivos, o agregar uno nuevo, fuerza la relectura. Y como la
        # lista de rutas es parte de la clave, DOS TAMBOS CON CATÁLOGOS DISTINTOS
        # no se pisan entre sí — con un caché de un solo slot, el segundo tambo
        # en pedir el catálogo se lo servía al primero.
        mtime = tuple((r, os.path.getmtime(r)) for r in presentes)
        if _cache.get(mtime) is not None:
            return _cache[mtime]
        crudos, errores = {}, []
        for ruta in presentes:
            try:
                # El primero que trae un toro gana: los archivos van de más
                # confiable a menos (el real le gana al simulado).
                for clave, datos_toro in _leer_excel(ruta).items():
                    crudos.setdefault(clave, datos_toro)
            except ImportError:
                return {"toros": {}, "por_nombre": {}, "error":
                        "Falta la librería openpyxl para leer el catálogo de toros "
                        "(pip install openpyxl)."}
            except Exception as exc:  # noqa: BLE001
                errores.append(f"{os.path.basename(ruta)}: {exc}")

        # Percentil de cada rasgo DENTRO del catálogo, con su dirección.
        valores = {clave: sorted(t["rasgos"][clave] for t in crudos.values()
                                 if clave in t["rasgos"])
                   for clave, _ in RASGOS.values()}
        for t in crudos.values():
            pcts = {}
            for col, (clave, alto_mejor) in RASGOS.items():
                if clave not in t["rasgos"]:
                    continue
                p = _percentil(t["rasgos"][clave], valores[clave])
                # El riesgo crece cuando el rasgo empeora: si más alto es
                # mejor, se invierte el percentil.
                pcts[clave] = (100.0 - p) if alto_mejor else p
            t["percentiles"] = pcts
            # Promedio de los rasgos DISPONIBLES (los que faltan se excluyen,
            # no se cuentan como 0 — mismo criterio que el resto de la app).
            t["riesgo"] = round(sum(pcts.values()) / len(pcts), 1) if pcts else None
            t["rasgos_usados"] = sorted(pcts)
            if not pcts:
                t["motivo"] = "El toro está en el catálogo pero sin ningún rasgo de salud cargado."

        datos = {
            "toros": crudos,
            # Índice por nombre en mayúsculas, para cruzar cuando la base
            # guarda el nombre del padre en vez del código NAAB.
            "por_nombre": {t["nombre"].upper(): clave for clave, t in crudos.items()
                           if t.get("nombre")},
            "error": "; ".join(errores) if errores else None,
            "archivos": [os.path.basename(r) for r in presentes],
            "simulados": sum(1 for t in crudos.values() if t.get("simulado")),
            "escala": (f"Percentil entre los {len(crudos)} toros del catálogo "
                       "(100 = el peor para salud, 0 = el mejor). Al agregar "
                       "toros los valores se reescalan."),
        }
        # Se guarda por clave de archivos. El diccionario se limpia si crece:
        # cada entrada es el catálogo entero y no hay motivo para tener más de
        # un par (una por tambo con catálogo propio).
        if len(_cache) > 8:
            _cache.clear()
        _cache[mtime] = datos
        return datos


def de_toro(identificador, rutas=None) -> dict | None:
    """Busca un toro por código NAAB o por nombre (sin distinguir mayúsculas).
    None = no está en el catálogo."""
    if not identificador:
        return None
    cat = catalogo(rutas)
    clave = str(identificador).strip().upper()
    if clave in cat["toros"]:
        return cat["toros"][clave]
    otra = cat["por_nombre"].get(clave)
    return cat["toros"].get(otra) if otra else None


def buscador(rutas=None):
    """`de_toro` con las rutas ya fijadas, para pasarlo como `padre_fn`.

    `herencia.de()` y `herencia.arbol()` esperan una función de UN argumento, y
    el catálogo a usar depende del tambo. Sin esto cada llamador armaba su
    propio lambda y era fácil olvidarse las rutas en uno y quedar leyendo el
    catálogo de otro tambo sin que nada avise.
    """
    return lambda ident: de_toro(ident, rutas)


def resumen(rutas=None) -> dict:
    """Para la pantalla de configuración/diagnóstico: cuántos toros tiene el
    catálogo, cuántos quedaron con riesgo calculable, y —importante— cuántos
    salen de datos SIMULADOS, para que la pantalla lo avise en vez de mostrar
    un riesgo inventado como si fuera medido."""
    cat = catalogo(rutas)
    toros = cat["toros"]
    con_riesgo = [t for t in toros.values() if t.get("riesgo") is not None]
    simulados = cat.get("simulados", 0)
    return {
        "error": cat.get("error"),
        "toros": len(toros),
        "con_riesgo": len(con_riesgo),
        "sin_rasgos": len(toros) - len(con_riesgo),
        "simulados": simulados,
        "hay_simulados": bool(simulados),
        "aviso_simulado": ("Hay datos genéticos FICTICIOS en uso "
                           f"({simulados} de {len(toros)} toros): sirven para probar el "
                           "cálculo, no para decidir nada sobre el rodeo.") if simulados else None,
        "archivos": cat.get("archivos", []),
        "escala": cat.get("escala"),
        "rasgos": [clave for clave, _ in RASGOS.values()],
    }
